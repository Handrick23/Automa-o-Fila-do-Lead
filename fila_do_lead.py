from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import customtkinter as ctk
import pandas as pd
import math
import os
import tkinter.filedialog as filedialog
from tkinter import messagebox

# =================================================================
# VARIÁVEIS GLOBAIS
# =================================================================
df_semana = None
df_consultores = None
df_mensal = None

# =================================================================
# FUNÇÕES AUXILIARES
# =================================================================
def aplicar_estilo_padrao(ws, font_header, fill_header, align_center):
    """Padroniza a estética das abas de base (Cabeçalho azul, R$ e larguras)"""
    for col_idx, column_cells in enumerate(ws.columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 25
        for row_idx, cell in enumerate(column_cells, 1):
            if row_idx == 1:
                cell.font, cell.fill, cell.alignment = font_header, fill_header, align_center
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                header_name = str(ws.cell(row=1, column=col_idx).value).upper()
                if any(x in header_name for x in ["VENDA", "TOTAL", "VALOR"]):
                    cell.number_format = '"R$ "#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

def processar_vendas(dataframe, sufixo=""):
    if dataframe is None or dataframe.empty:
        return pd.DataFrame(columns=['Consultor', f'Venda Novo{sufixo}', f'Venda Existente{sufixo}', f'Total{sufixo}'])
    dataframe.columns = [str(c).strip().title() for c in dataframe.columns]
    col_valor = 'Venda'
    if col_valor not in dataframe.columns:
        cols_num = dataframe.select_dtypes(include=['number']).columns
        col_valor = cols_num[0] if len(cols_num) > 0 else 'Venda'
    resumo = dataframe.pivot_table(index='Consultor', columns='Tipo Cliente', values=col_valor, aggfunc='sum').fillna(0).reset_index()
    for col in ['Novo', 'Existente']:
        if col not in resumo.columns: resumo[col] = 0
    resumo = resumo.rename(columns={'Novo': f'Venda Novo{sufixo}', 'Existente': f'Venda Existente{sufixo}'})
    resumo[f'Total{sufixo}'] = resumo[f'Venda Novo{sufixo}'] + resumo[f'Venda Existente{sufixo}']
    return resumo

def gerar_fila_do_lead():
    global df_semana, df_consultores, df_mensal
    if df_semana is None or df_consultores is None:
        messagebox.showwarning("Aviso", "Por favor, carregue o arquivo Excel primeiro.")
        return

    vendas_semana = processar_vendas(df_semana)
    vendas_mensal = processar_vendas(df_mensal, sufixo="_Mensal")

    df_base_geral = pd.merge(df_consultores[['Consultor', 'Equipe', 'Justificativa']], vendas_semana, on='Consultor', how='left').fillna(0)
    df_base_geral = pd.merge(df_base_geral, vendas_mensal, on='Consultor', how='left').fillna(0)
    df_base_geral['Férias?'] = df_base_geral['Justificativa'].apply(lambda x: 'Sim' if str(x).upper() == 'FÉRIAS' else 'Não')
    
    base_fila = df_base_geral[df_base_geral['Férias?'] == 'Não'].copy()
    
    def mapear_unidade_comercial(equipe):
        e = str(equipe).upper().strip()
        mapeamento = {
            'GRANDES CONTAS SP': 'SPO', 'SP 1': 'SPO', 'SPO': 'SPO', 'SP': 'SPO',
            'GRANDES CONTAS BH': 'BHZ', 'MG 1': 'BHZ', 'BHZ': 'BHZ', 'BH': 'BHZ',
            'GRANDES CONTAS RJ': 'RJO', 'RJ 1': 'RJO', 'RJO': 'RJO', 'RJ': 'RJO',
            'GRANDES CONTAS CTA': 'CTA', 'CURITIBA': 'CTA', 'CTA': 'CTA',
            'SP INTERIOR': 'SP INTERIOR', 'SPI': 'SP INTERIOR'
        }
        return mapeamento.get(e, e)
    base_fila['Filial_Final'] = base_fila['Equipe'].apply(mapear_unidade_comercial)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fila do Lead"
    
    # Estilos
    side_m = Side(style='medium', color='000000')
    side_t = Side(style='thin', color='000000')
    fill_header = PatternFill(start_color="384B59", end_color="384B59", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_red = Font(bold=True, color="FF0000", size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D']: ws.column_dimensions[col].width = 15

    row_idx = 1
    for filial in sorted(base_fila['Filial_Final'].unique()):
        grupo = base_fila[base_fila['Filial_Final'] == filial].copy()

        # RESTAURAÇÃO DA REGRA DE NEGÓCIO ORIGINAL (Fila 1 = 50% melhores)
        cat_a = grupo[grupo['Total'] > 0].sort_values(by=['Venda Novo', 'Total'], ascending=False)
        n_f1 = math.ceil(len(cat_a) / 2) if len(cat_a) > 0 else 0
        lista_f1 = cat_a.head(n_f1)
        
        sobras_cat_a = cat_a.tail(len(cat_a) - n_f1).sort_values(by='Total', ascending=False)
        cat_b = grupo[(grupo['Total'] == 0) & (grupo['Total_Mensal'] > 0)].sort_values(by='Total_Mensal', ascending=False)
        cat_c = grupo[(grupo['Total'] == 0) & (grupo['Total_Mensal'] == 0)].sample(frac=1)
        lista_f2 = pd.concat([sobras_cat_a, cat_b, cat_c])

        # Escrita Cabeçalho Filial
        cell = ws.cell(row=row_idx, column=1, value=f"{filial} Comercial")
        cell.fill, cell.font, cell.alignment, cell.border = fill_header, font_header, align_center, Border(top=side_m, left=side_m, right=side_m)
        for col, txt in enumerate(["NOVO", "EXISTENTE", "TOTAL"], 2):
            c = ws.cell(row=row_idx, column=col, value=txt)
            c.font, c.alignment, c.border = Font(bold=True, size=9), align_center, Border(top=side_m, bottom=side_t)
        row_idx += 1

        # Fila 1
        ws.cell(row=row_idx, column=1, value="Fila 1").font = font_red
        ws.cell(row=row_idx, column=1).alignment, ws.cell(row=row_idx, column=1).border = align_center, Border(left=side_m, right=side_m)
        row_idx += 1
        for _, r in lista_f1.iterrows():
            ws.cell(row=row_idx, column=1, value=str(r['Consultor']).title()).alignment = align_center
            ws.cell(row=row_idx, column=1).border = Border(left=side_m, right=side_m)
            for col, val in enumerate([r['Venda Novo'], r['Venda Existente'], r['Total']], 2):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.number_format, c.alignment = '"R$ "#,##0.00', align_right
            row_idx += 1

        # Fila 2
        for _ in range(1): ws.cell(row=row_idx, column=1).border = Border(left=side_m, right=side_m); row_idx += 1
        ws.cell(row=row_idx, column=1, value="Fila 2").font = font_red
        ws.cell(row=row_idx, column=1).alignment, ws.cell(row=row_idx, column=1).border = align_center, Border(left=side_m, right=side_m)
        row_idx += 1
        for i, (_, r) in enumerate(lista_f2.iterrows()):
            ws.cell(row=row_idx, column=1, value=str(r['Consultor']).title()).alignment = align_center
            borda = Border(left=side_m, right=side_m, bottom=side_m) if i == len(lista_f2)-1 else Border(left=side_m, right=side_m)
            ws.cell(row=row_idx, column=1).border = borda
            for col, val in enumerate([r['Venda Novo'], r['Venda Existente'], r['Total']], 2):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.number_format, c.alignment = '"R$ "#,##0.00', align_right
                if i == len(lista_f2)-1: c.border = Border(bottom=side_m)
            row_idx += 1
        row_idx += 2

    # Gerar e Formatar as outras abas conforme solicitado
    abas = {"Base Semanal": df_semana, "Base Mensal": df_mensal, "Base Consultores": df_consultores, "Resumo Processamento": df_base_geral}
    for nome, df in abas.items():
        if df is not None:
            nws = wb.create_sheet(title=nome)
            for r in dataframe_to_rows(df, index=False, header=True): nws.append(r)
            aplicar_estilo_padrao(nws, font_header, fill_header, align_center)

    nome_final = "Fila_do_Lead.xlsx"
    wb.save(nome_final)
    messagebox.showinfo("Sucesso", "Fila e Base geradas com sucesso!")
    os.startfile(nome_final)

def importar_planilha():
    """
    UI LOGIC: Abre o seletor de arquivos e carrega as abas específicas.
    """
    global df_semana, df_consultores, df_mensal
    caminho = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if caminho:
        try:
            xls = pd.ExcelFile(caminho)
            abas_map = {a.strip().upper(): a for a in xls.sheet_names}
            
            aba_sem = abas_map.get('BASE LEAD') or abas_map.get('BASE SEMANAL')
            aba_men = abas_map.get('BASE MENSAL')
            aba_con = abas_map.get('CONSULTORES')

            if aba_sem and aba_men and aba_con:
                df_semana = pd.read_excel(xls, aba_sem)
                df_mensal = pd.read_excel(xls, aba_men)
                df_consultores = pd.read_excel(xls, aba_con)
                
                for d in [df_semana, df_mensal, df_consultores]:
                    d.columns = [str(c).strip().title() for c in d.columns]
                    if 'Consultor' in d.columns:
                        d['Consultor'] = d['Consultor'].astype(str).str.strip().str.upper()
                messagebox.showinfo("Sucesso", "Bases carregadas com sucesso!")
            else:
                messagebox.showerror("Erro", "Abas obrigatórias não encontradas!")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar arquivo: {e}")

# =================================================================
# INTERFACE GRÁFICA (CustomTkinter)
# =================================================================
ctk.set_appearance_mode("dark")
app = ctk.CTk()
app.title("Fila Lead Pro 2026")
app.geometry('400x280')

ctk.CTkLabel(app, text="Fila do Lead", font=("Arial", 24, "bold")).pack(pady=30)

f_main = ctk.CTkFrame(app)
f_main.pack(pady=10, padx=30, fill="both", expand=True)

ctk.CTkButton(f_main, text="Carregar Planilha Excel", command=importar_planilha, width=250).pack(pady=20)
ctk.CTkButton(f_main, text="Gerar Fila Ranqueada", command=gerar_fila_do_lead, fg_color="#27ae60", width=250).pack(pady=10)

app.mainloop()
